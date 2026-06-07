"""
Apply the Clempo brand charter to the Décideurs hospitaliers Google Sheet export.

Mirrors the journalistes lead magnet:
  - Adds a "00 · couverture" cover sheet (wordmark, hero, stats, bio, CTAs)
  - Restyles the data sheet with editorial headers and alternating row fills
  - Hides gridlines, freezes panes under the headers

Input:  /tmp/decideurs.xlsx (download the Google Sheet as .xlsx)
Output: /tmp/decideurs-clempo.xlsx
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter

SRC = '/tmp/decideurs.xlsx'
DST = '/tmp/decideurs-clempo.xlsx'

# ── Brand tokens (mirror the journalistes cover) ──
PAPER       = 'FFEDEBE4'   # warm off-white background
INK         = 'FF0A0A0B'   # primary text
STEEL       = 'FF6B6F7A'   # secondary text
GRAPHITE    = 'FF2A2D35'   # body emphasis
SIGNAL      = 'FF00D68F'   # accent green (CTA backgrounds)
SIGNAL_DEEP = 'FF009E68'   # accent green (text on paper)
WHITE       = 'FFFFFFFF'
ROW_ZEBRA   = 'FFF4F4F2'

INTER = 'Inter'
MONO  = 'JetBrains Mono'
SERIF = 'Instrument Serif'


def paint_cover(ws):
    """Write the cover sheet on a blank worksheet — mirrors journalistes cover."""
    ws.title = '00 · couverture'
    ws.sheet_view.showGridLines = False

    # Column widths (A=narrow margin, B-L=content, A repeats narrowly so layout is consistent)
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # Row heights
    row_heights = {
         1: 19.5,  2: 19.5,  3: 27.8,  4: 19.5,  5: 19.5,  6: 19.5,  7: 19.5,
         8: 19.5,  9: 19.5, 10: 19.5, 11: 49.5, 12: 19.5, 13: 49.5, 14: 19.5,
        15: 19.5, 16: 19.5, 17: 19.5, 18: 19.5, 19: 19.5, 20: 19.5, 21: 13.5,
        22: 30.0, 23: 15.8, 24: 19.5, 25: 19.5, 26:  1.5, 27: 19.5, 28: 19.5,
        29: 19.5, 30: 25.5, 31: 19.5, 32: 19.5, 33: 19.5, 34: 19.5, 35: 19.5,
        36: 19.5, 37: 19.5, 38: 13.5, 39: 24.0, 40: 21.8, 41: 13.5, 42: 19.5,
        43: 19.5, 44: 19.5, 45:  6.0, 46: 19.5, 47: 19.5,
    }
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    # Paint background on every cell of the cover area
    for r in range(1, 48):
        for c in range(1, 15):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=PAPER)

    def put(coord, value, font_name, size, bold, color, align='left', bg=PAPER):
        cell = ws[coord]
        cell.value = value
        cell.font = Font(name=font_name, size=size, bold=bold, color=color)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')

    # ── HEADER ──
    put('B3',  'clempo.',                              INTER, 22, True,  INK,         'left')
    put('L3',  'HOSPITAL DATABASE · VOL. 02',          MONO,   8, False, STEEL,       'right')
    put('L4',  'ÉDITION 2026',                         MONO,   8, False, STEEL,       'right')

    # ── EYEBROW ──
    put('B9',  '— DÉCIDEURS HOSPITALIERS · FRANCE',    MONO,  10, True,  SIGNAL_DEEP, 'left')

    # ── HERO ──
    put('B11', 'Décideurs hospitaliers.',              INTER, 36, True,  INK,         'left')
    put('B13', 'France.',                              INTER, 36, True,  INK,         'left')

    # ── DESCRIPTION ──
    put('B16', '8 836 décideurs santé, 1 849 établissements, prêts à filtrer.',
        SERIF, 14, False, GRAPHITE, 'left')
    put('B17', 'Base de travail pour cold outreach, ABM hôpital et mapping commercial.',
        SERIF, 14, False, STEEL,    'left')

    # ── STATS GRID ──
    put('B21', '// 01 — PUBLIC',                       MONO,   8, False, STEEL,       'left')
    put('F21', '// 02 — PRIVÉ & ESPIC',                MONO,   8, False, STEEL,       'left')
    put('J21', '// 03 — TOTAL',                        MONO,   8, False, STEEL,       'left')
    put('B22', 7299,                                   INTER, 26, True,  INK,         'left')
    put('F22', 1410,                                   INTER, 26, True,  INK,         'left')
    put('J22', 8836,                                   INTER, 26, True,  SIGNAL_DEEP, 'left')
    put('B23', 'CH · CHU · CHRU · HL · EHPAD · CHS',   INTER, 10, False, STEEL,       'left')
    put('F23', 'cliniques · ESPIC · CLCC · privé',     INTER, 10, False, STEEL,       'left')
    put('J23', '1 849 établissements · 1 287 villes',  INTER, 10, False, STEEL,       'left')

    # ── BIO ──
    put('B28', '— QUI EST DERRIÈRE CETTE BASE',        MONO,   9, True,  SIGNAL_DEEP, 'left')
    put('B30', 'Clément Pouget-Osmont — Fractional CMO santé.', INTER, 18, True, INK, 'left')
    put('B32', "12 ans dans la santé, dont 5 chez Doctolib. J'accompagne fondateurs healthtech",
        INTER, 11, False, GRAPHITE, 'left')
    put('B33', 'et directions marketing pharma sur le positionnement, le funnel, et le GTM.',
        INTER, 11, False, GRAPHITE, 'left')
    put('B34', 'Missions fractional, 2 à 4 jours par semaine. Démarrage sous 15 jours.',
        INTER, 11, False, STEEL,    'left')

    # ── CTAs ──
    put('B39', "// PARLER D'UNE MISSION", MONO,  9, True, SIGNAL,      'left', bg=INK)
    put('H39', '// GUIDE D\'USAGE',       MONO,  9, True, INK,         'left', bg=SIGNAL)
    put('B40', 'clempo.fr/booking →',     INTER, 16, True, WHITE,      'left', bg=INK)
    put('H40', 'Comment utiliser cette base →', INTER, 13, True, INK,  'left', bg=SIGNAL)
    put('H41', 'clempo.fr/decideurs-hospitaliers', MONO, 9, False, INK,'left', bg=SIGNAL)
    # Merge CTA cells
    ws.merge_cells('B40:D40')
    ws.merge_cells('H40:L40')
    ws.merge_cells('H41:L41')
    # Paint merge backgrounds on every merged cell
    for coord in ('C40', 'D40', 'I40', 'J40', 'K40', 'L40'):
        ws[coord].fill = PatternFill('solid', fgColor=INK if coord in ('C40','D40') else SIGNAL)
    for coord in ('I41', 'J41', 'K41', 'L41'):
        ws[coord].fill = PatternFill('solid', fgColor=SIGNAL)

    # ── FOOTER ──
    put('B47', '// CLEMPO.FR',                          MONO,  9, True,  SIGNAL_DEEP, 'left')
    put('L47', 'DOCUMENT DE TRAVAIL · TOUS DROITS RÉSERVÉS', MONO, 8, False, STEEL,   'right')


def restyle_data_sheet(ws):
    """Apply Clempo editorial styling to the data sheet."""
    ws.title = '01 · décideurs hospitaliers'
    ws.sheet_view.showGridLines = False

    n_rows = ws.max_row
    n_cols = ws.max_column

    # Column widths tuned for the 19 columns of the décideurs dataset
    widths = {
        'A': 14, 'B': 16, 'C': 30, 'D': 40, 'E': 32, 'F': 16, 'G': 38,
        'H': 26, 'I': 16, 'J': 32, 'K': 28, 'L': 40, 'M': 16, 'N': 20,
        'O': 10, 'P': 12, 'Q': 14, 'R': 12, 'S': 24,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Insert 2 rows at top so we can add the section banner + spacer
    ws.insert_rows(1, amount=2)
    ws.row_dimensions[1].height = 18.0
    ws.row_dimensions[2].height = 7.5
    ws.row_dimensions[3].height = 30.0

    # Row 1 — section banner (merged across all columns)
    banner = ws.cell(row=1, column=1)
    banner.value = f'// 01 — DÉCIDEURS HOSPITALIERS · {n_rows - 1} CONTACTS · FRANCE'
    banner.font = Font(name=MONO, size=9, bold=True, color=SIGNAL_DEEP)
    banner.alignment = Alignment(horizontal='left', vertical='center')
    last_col_letter = get_column_letter(n_cols)
    ws.merge_cells(f'A1:{last_col_letter}1')

    # Row 3 — column headers (Inter 10pt bold, white on ink)
    header_font = Font(name=INTER, size=10, bold=True, color=WHITE)
    header_fill = PatternFill('solid', fgColor=INK)
    header_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=3, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Rows 4+ — body (Inter 10pt + alternating zebra fill)
    body_font = Font(name=INTER, size=10, bold=False, color=INK)
    zebra_fill = PatternFill('solid', fgColor=ROW_ZEBRA)
    body_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    for r in range(4, n_rows + 3):  # n_rows already accounted for original, +2 inserted
        is_zebra = (r % 2 == 1)
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.alignment = body_align
            if is_zebra:
                cell.fill = zebra_fill
        ws.row_dimensions[r].height = 21.75

    # Freeze under the headers so they stay visible on scroll
    ws.freeze_panes = 'A4'


def main():
    wb = load_workbook(SRC)

    # Restyle the existing data sheet (in place)
    data_sheet = wb[wb.sheetnames[0]]
    restyle_data_sheet(data_sheet)

    # Create the cover sheet at index 0
    cover = wb.create_sheet(title='00 · couverture', index=0)
    paint_cover(cover)

    wb.save(DST)
    print(f'✓ Wrote {DST}')
    print(f'  Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    main()
