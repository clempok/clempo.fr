#!/usr/bin/env python3
"""Mise à jour trimestrielle des parts de télétransmission GIE SESAM-Vitale.

Récupère le dump JSON de l'API du GIE et met à jour, pour les 14 spécialités
publiées sur clempo.fr :
  - public/data/specialites/<slug>.csv   (lignes brutes, splice additif)
  - public/data/specialites/<slug>.json  (données BarChartRace, splice additif)
  - public/data/specialites/<slug>.xlsx  (classeur téléchargeable, régénéré)
  - Extract GIE/<name>_bar_chart_race.html (blob DATA remplacé)

Usage :
  python3 scripts/update_gie_teletransmission.py            # télécharge l'API
  python3 scripts/update_gie_teletransmission.py dump.json  # dump déjà téléchargé

Le splice est strictement additif : les mois déjà publiés sont vérifiés
inchangés (le script s'arrête si le GIE a révisé l'historique).
"""
import csv
import io
import json
import re
import sys
import unicodedata
import urllib.request
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / 'public' / 'data' / 'specialites'
EXTRACT_DIR = ROOT / 'Extract GIE'

API_URL = (
    'https://www.sesam-vitale.fr:443/web/sesam-vitale/parts-de-teletransmission'
    '?p_p_id=fr_sesamvitale_portail_chiffrespdm_web_portlet_chiffresPDMPortlet'
    '&p_p_lifecycle=2&p_p_state=normal&p_p_mode=view'
    '&p_p_resource_id=%2Fchiffrespdm%2FgetJsonResponse&p_p_cacheability=cacheLevelPage'
)

# catPro GIE -> (slug site, préfixe fichier "Extract GIE")
CATEGORIES = {
    'Médecins généralistes': ('medecins-generalistes', 'medecins'),
    'Médecins spécialistes': ('medecins-specialistes', 'medecins_specialistes'),
    'Dentistes': ('dentistes', 'dentistes'),
    'Infirmiers': ('infirmiers', 'infirmiers'),
    'Masseurs kinésithérapeutes': ('kines', 'kines'),
    'Orthophonistes': ('orthophonistes', 'orthophonistes'),
    'Orthoptistes': ('orthoptistes', 'orthoptistes'),
    'Pharmacies': ('pharmacies', 'pharmacies'),
    'Opticiens': ('opticiens', 'opticiens'),
    'Audioprothèsistes': ('audioprothesistes', 'audioprothesistes'),
    'Sages-femmes': ('sages-femmes', 'sages_femmes'),
    'Pédicures-Podologues': ('pedicures-podologues', 'podologues'),
    'Laboratoires': ('laboratoires-analyses', 'laboratoires'),
    'Centres de santé': ('centres-sante', 'centres_sante'),
}

AUTRES_RAW = 'Autres éditeurs/solutions non répertoriés'
AUTRES_DISPLAY = 'Autres (non répertoriés)'

# Noms commerciaux d'usage (progiciel GIE -> nom affiché)
RENAME_PROG = {
    'AXIAM': 'AxiSanté (AxiAM)',
    'EXPRESS VITALE': 'Medistory (Express Vitale)',
    'VITALZEN': 'Weda (VitalZen)',
    'AGATHE YOU - E.MOTION': 'Agathe YOU (E.Motion)',
    'LGPI': 'ID. (LGPI)',
}

# Libellés éditeur modifiés rétroactivement par le GIE : on les ramène au
# libellé historique publié, sinon le splice casse la continuité des clés.
EDITEUR_ALIASES = {
    'JULIE': 'JULIE SOLUTIONS',  # renommé par le GIE en 2026 (rachat Imagex)
}

MONTH_NAMES = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
               'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

THRESHOLD = 0.01  # part (%) non arrondie minimale pour figurer dans monthly_data


def display_key(prog, edit):
    if prog == AUTRES_RAW:
        return AUTRES_DISPLAY
    if prog in RENAME_PROG:
        return f"{RENAME_PROG[prog]} — {edit}"
    return f"{prog} — {edit}"


def display_prog_short(prog):
    if prog == 'total':
        return 'total'
    return RENAME_PROG.get(prog, prog)


def month_fmt(d):  # 201901 -> '2019-01'
    d = str(d)
    return f"{d[:4]}-{d[4:6]}"


def csv_sort_key(r):
    # ordre observé : groupe/éditeur/progiciel insensibles à la casse,
    # '__NO_GROUP__*' relégué en fin ('_' remplacé par '~')
    return tuple(r[k].lower().replace('_', '~') for k in ('groupe', 'editeur', 'progiciel'))


def load_dump(path=None):
    if path:
        raw = json.load(open(path))
    else:
        print(f"Téléchargement de l'API GIE…")
        with urllib.request.urlopen(API_URL, timeout=120) as resp:
            raw = json.load(resp)
    return raw['chiffres'] if isinstance(raw, dict) else raw


def update_csv(path, cat_rows, old_months, new_months):
    """Vérifie l'existant et ajoute les nouveaux mois. Retourne les lignes complètes."""
    existing = open(path, encoding='utf-8').read()
    old_rows = list(csv.DictReader(io.StringIO(existing), delimiter=';'))
    old_set = {(r['dateData'], r['groupe'], r['editeur'], r['progiciel'], r['nbPsFacturation'])
               for r in old_rows}
    dump_old = {(str(r['dateData']), r['groupe'], r['editeur'], r['progiciel'], str(r['nbPsFacturation']))
                for r in cat_rows if str(r['dateData']) in old_months}
    if old_set != dump_old:
        raise SystemExit(f"⚠ {path.name} : l'historique GIE a été révisé — splice impossible, "
                         f"({len(old_set - dump_old)} lignes locales absentes du dump, "
                         f"{len(dump_old - old_set)} lignes du dump absentes en local)")
    lines = [existing]
    for m in new_months:
        month_rows = sorted(
            ({'groupe': r['groupe'], 'editeur': r['editeur'], 'progiciel': r['progiciel'],
              'nbPsFacturation': str(r['nbPsFacturation'])}
             for r in cat_rows if str(r['dateData']) == m),
            key=csv_sort_key)
        for r in month_rows:
            lines.append(f"\n{m};{r['groupe']};{r['editeur']};{r['progiciel']};{r['nbPsFacturation']}")
    open(path, 'w', encoding='utf-8').write(''.join(lines))
    return list(csv.DictReader(io.StringIO(''.join(lines)), delimiter=';'))


def month_shares(rows_month, total):
    """{clé affichée: part non arrondie} pour un mois (hors ligne total)."""
    agg = defaultdict(int)
    for r in rows_month:
        agg[display_key(r['progiciel'], r['editeur'])] += int(r['nbPsFacturation'])
    return {k: 100 * n / total for k, n in agg.items()}


def update_json(path, all_rows, new_months):
    J = json.load(open(path, encoding='utf-8'))
    bymonth = defaultdict(list)
    totals = {}
    for r in all_rows:
        if r['progiciel'] == 'total':
            totals[r['dateData']] = int(r['nbPsFacturation'])
        else:
            bymonth[r['dateData']].append(r)
    for m in new_months:
        shares = month_shares(bymonth[m], totals[m])
        kept = {k: round(v, 4) for k, v in shares.items() if v > THRESHOLD}
        J['months'].append(month_fmt(m))
        J['monthly_data'][month_fmt(m)] = kept
        ranked = sorted(((k, round(v, 4)) for k, v in shares.items()),
                        key=lambda x: (-x[1], x[0]))
        J['top10'].append([[k, v] for k, v in ranked[:10]])
    J['all_progiciels'] = sorted(set(J['all_progiciels'])
                                 | {k for m in new_months for k in J['monthly_data'][month_fmt(m)]})
    out = json.dumps(J, ensure_ascii=False, separators=(',', ':'))
    open(path, 'w', encoding='utf-8').write(out)
    return J


def update_html(path, J):
    if not path.exists():
        print(f"  (pas de {path.name})")
        return
    html = open(path, encoding='utf-8').read()
    data = {k: J[k] for k in ('months', 'monthly_data', 'all_progiciels', 'top10')}
    blob = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    new_html, n = re.subn(r'const DATA = \{.*?\};', lambda _: f'const DATA = {blob};',
                          html, count=1, flags=re.S)
    if n != 1:
        raise SystemExit(f"⚠ {path.name} : blob DATA introuvable")
    open(path, 'w', encoding='utf-8').write(new_html)


def build_xlsx(path, all_rows, cat_name, extraction_date):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='305496')

    months = sorted({r['dateData'] for r in all_rows})
    totals = {r['dateData']: int(r['nbPsFacturation']) for r in all_rows if r['progiciel'] == 'total'}
    details = [r for r in all_rows if r['progiciel'] != 'total']

    wb = openpyxl.Workbook()

    # ---- Info
    ws = wb.active
    ws.title = 'Info'
    editeurs = sorted({r['editeur'] for r in details})
    groupes = {r['groupe'] for r in details
               if not r['groupe'].startswith('__NO_GROUP__')}
    progiciels = {(r['progiciel'], r['editeur']) for r in details}
    info_rows = [
        (f"Parts de télétransmission - Catégorie {cat_name.upper()}",),
        (),
        ('Source', 'GIE SESAM-Vitale'),
        ('URL', 'https://www.sesam-vitale.fr/parts-de-teletransmission'),
        ('Catégorie', cat_name),
        ('Période couverte', f"{month_fmt(months[0])} → {month_fmt(months[-1])}"),
        ('Nombre de mois', len(months)),
        ('Nombre total de lignes', len(all_rows)),
        ('Nombre de progiciels distincts', len(progiciels)),
        ("Nombre d'éditeurs distincts", len(editeurs)),
        ('Nombre de groupes distincts', len(groupes)),
        ("Date d'extraction", extraction_date),
        (),
        ('Onglets',),
        ('Données', 'Lignes brutes + part calculée'),
        ('Pivot Éditeurs %', 'Tableau croisé : éditeur × mois → %'),
        ('Pivot Éditeurs Nb', 'Tableau croisé : éditeur × mois → nb facturants'),
        (),
        ('Notes',),
        ('• Lignes "TOTAL" : ensemble des facturants du mois.',),
        ('• Part % calculée = Nombre de facturants / Total mois.',),
        ("• Noms commerciaux d'usage utilisés quand différents du nom GIE :",),
    ] + [(f"   {raw} → {new}",) for raw, new in RENAME_PROG.items()]
    for row in info_rows:
        ws.append(row)
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 55

    # ---- Données
    ws = wb.create_sheet('Données')
    header = ['Période', 'Année', 'Mois (n°)', 'Mois', 'Groupe', 'Éditeur', 'Progiciel',
              'Nombre de facturants', 'Total mois', 'Part %', 'Type']
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
    for m in months:
        yyyy, mm = int(str(m)[:4]), int(str(m)[4:6])
        month_rows = sorted((r for r in details if r['dateData'] == m),
                            key=lambda r: (-int(r['nbPsFacturation']), r['editeur'], r['progiciel']))
        for r in month_rows:
            groupe = r['groupe']
            if groupe.startswith('__NO_GROUP__') or groupe == AUTRES_RAW:
                groupe = None
            ws.append([month_fmt(m), yyyy, mm, MONTH_NAMES[mm - 1], groupe, r['editeur'],
                       display_prog_short(r['progiciel']), int(r['nbPsFacturation']),
                       totals[m], int(r['nbPsFacturation']) / totals[m], 'DETAIL'])
            ws.cell(row=ws.max_row, column=10).number_format = '0.00%'
        ws.append([month_fmt(m), yyyy, mm, MONTH_NAMES[mm - 1], None, 'total', 'total',
                   totals[m], totals[m], 1, 'TOTAL'])
        ws.cell(row=ws.max_row, column=10).number_format = '0.00%'
    for col, w in zip('ABCDEFGHIJK', (10, 8, 10, 12, 35, 35, 35, 16, 12, 10, 10)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'

    # ---- Pivots
    def sortkey(e):
        return unicodedata.normalize('NFD', e).encode('ascii', 'ignore').decode().lower()
    ed_sorted = sorted(editeurs, key=sortkey)
    by_ed_month = defaultdict(int)
    for r in details:
        by_ed_month[(r['editeur'], r['dateData'])] += int(r['nbPsFacturation'])

    for sheet_name, as_pct in (('Pivot Éditeurs %', True), ('Pivot Éditeurs Nb', False)):
        ws = wb.create_sheet(sheet_name)
        ws.append(['Éditeur'] + [month_fmt(m) for m in months])
        for c in range(1, len(months) + 2):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
        for e in ed_sorted:
            vals = []
            for m in months:
                n = by_ed_month.get((e, m), 0)
                vals.append(n / totals[m] if as_pct else n)
            ws.append([e] + vals)
            if as_pct:
                for c in range(2, len(months) + 2):
                    ws.cell(row=ws.max_row, column=c).number_format = '0.00%'
        ws.column_dimensions['A'].width = 45
        for c in range(2, len(months) + 2):
            ws.column_dimensions[get_column_letter(c)].width = 10
        ws.freeze_panes = 'B2'

    wb.save(path)


def main():
    dump_path = sys.argv[1] if len(sys.argv) > 1 else None
    extraction_date = sys.argv[2] if len(sys.argv) > 2 else None
    if not extraction_date:
        raise SystemExit("Usage: update_gie_teletransmission.py [dump.json] '21 juillet 2026'")
    rows = load_dump(dump_path)
    for r in rows:
        r['dateData'] = str(r['dateData'])
        if r['editeur'] in EDITEUR_ALIASES:
            old_label = EDITEUR_ALIASES[r['editeur']]
            if r['groupe'] == f"__NO_GROUP__{r['editeur']}":
                r['groupe'] = f"__NO_GROUP__{old_label}"
            r['editeur'] = old_label

    for cat_name, (slug, prefix) in CATEGORIES.items():
        cat_rows = [r for r in rows if r['catPro'] == cat_name]
        csv_path = DATA_DIR / f'{slug}.csv'
        J_old = json.load(open(DATA_DIR / f'{slug}.json', encoding='utf-8'))
        old_months = {m.replace('-', '') for m in J_old['months']}
        new_months = sorted({r['dateData'] for r in cat_rows} - old_months)
        print(f"— {slug}: +{len(new_months)} mois {[month_fmt(m) for m in new_months]}")
        if not new_months:
            continue
        all_rows = update_csv(csv_path, cat_rows, old_months, new_months)
        J = update_json(DATA_DIR / f'{slug}.json', all_rows, new_months)
        update_html(EXTRACT_DIR / f'{prefix}_bar_chart_race.html', J)
        build_xlsx(DATA_DIR / f'{slug}.xlsx', all_rows, cat_name, extraction_date)
    print('OK')


if __name__ == '__main__':
    main()
